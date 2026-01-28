#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
税负调整脚本
使用 formulas + openpyxl 实现纯 Python Excel 公式计算
"""

import os
import shutil
from datetime import datetime
import formulas
import openpyxl


class TaxAdjuster:
    """税负调整器 - 使用 formulas + openpyxl 实现"""

    # G25 安全范围
    G25_MIN = 0.85
    G25_MAX = 1.00

    # E18 安全范围（年利润总额）
    E18_MIN = 0
    E18_MAX = 10_000_000

    # 毛利率安全范围
    MARGIN_MIN = 0.70
    MARGIN_MAX = 0.90

    # B11 安全范围（产品成本加工费）
    B11_MIN = 0
    B11_MAX = 500_000

    # F20 安全范围（生产成本月结表）
    F20_MIN = -40_000
    F20_MAX = 40_000

    # H11 安全范围（生产成本月结表）
    H11_MIN = -10
    H11_MAX = 10

    # 毛利率单元格配置
    MARGIN_CELL = 'J14'
    MARGIN_SHEET = '生产成本月结表'

    def __init__(self, file_path, progress_callback=None):
        """初始化，保存文件路径

        Args:
            file_path: Excel 文件路径
            progress_callback: 进度回调函数，签名为 callback(progress, message)
                              progress: 0-100 的进度值
                              message: 进度描述文字
        """
        self.file_path = os.path.abspath(file_path)
        self._filename = os.path.basename(self.file_path)
        self.temp_file_path = None  # 临时副本文件路径
        self._model = None
        self._progress_callback = progress_callback

    def _report_progress(self, progress, message=""):
        """报告进度"""
        if self._progress_callback:
            self._progress_callback(progress, message)

    def _create_temp_copy(self):
        """创建原文件的临时副本"""
        if self.temp_file_path is None:
            base, ext = os.path.splitext(self.file_path)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            self.temp_file_path = f"{base}_temp_{timestamp}{ext}"
            shutil.copy2(self.file_path, self.temp_file_path)
        return self.temp_file_path

    def _cleanup_temp_file(self):
        """清理临时副本文件"""
        if self.temp_file_path and os.path.exists(self.temp_file_path):
            try:
                os.remove(self.temp_file_path)
            except Exception:
                pass  # 忽略清理失败
            self.temp_file_path = None

    def _cell_key(self, sheet_name, cell):
        """生成 formulas 单元格引用键"""
        return f"'[{self._filename}]{sheet_name}'!{cell}"

    def _load_model(self):
        """加载 formulas ExcelModel"""
        if self._model is None:
            temp_path = self._create_temp_copy()
            # 更新 _filename 为临时文件名，因为 formulas 使用文件名作为键的一部分
            self._filename = os.path.basename(temp_path)
            self._model = formulas.ExcelModel().loads(temp_path).finish()

    def _unload_model(self, save_to_original=False):
        """卸载模型并清理

        Args:
            save_to_original: 如果为 True，将副本内容复制回原文件
        """
        self._model = None

        # 如果需要保存到原文件，先复制再清理
        if save_to_original and self.temp_file_path:
            shutil.copy2(self.temp_file_path, self.file_path)

        # 清理临时文件
        self._cleanup_temp_file()

    def _get_value(self, solution, sheet_name, cell):
        """从 formulas solution 获取单元格值"""
        key = self._cell_key(sheet_name, cell)
        if key in solution:
            val = solution[key].value
            # formulas 返回 numpy 数组，需要提取标量值
            if hasattr(val, '__iter__') and not isinstance(val, str):
                try:
                    return val[0][0] if len(val) > 0 and hasattr(val[0], '__len__') else val[0]
                except (IndexError, TypeError):
                    return val
            return val
        return None

    def _calculate(self, inputs=None):
        """使用 formulas 计算，返回 solution"""
        if inputs is None:
            inputs = {}
        return self._model.calculate(inputs=inputs)

    def _to_number(self, value, default=0):
        """将值转换为数字"""
        if value is None:
            return default
        if isinstance(value, (int, float)):
            return value
        if isinstance(value, str):
            try:
                return float(value)
            except ValueError:
                return default
        return default

    def _check_range(self, value, min_val, max_val, name):
        """检查值是否在安全范围内，返回 (is_safe, message)"""
        if value < min_val:
            return False, f"{name} ({value:.4f}) 低于安全下限 ({min_val})"
        if value > max_val:
            return False, f"{name} ({value:.4f}) 超过安全上限 ({max_val})"
        return True, None

    def _check_margin_cell(self):
        """检查 J14 单元格是否有有效毛利率值

        Returns:
            (is_valid, margin_value, error_message)
        """
        try:
            wb = openpyxl.load_workbook(self.temp_file_path or self.file_path, data_only=True)
            ws = wb[self.MARGIN_SHEET]
            value = ws[self.MARGIN_CELL].value
            wb.close()

            if value is None:
                return False, None, f"请在 Excel 的 '{self.MARGIN_SHEET}' 工作表 {self.MARGIN_CELL} 单元格中设置毛利率值"

            margin = self._to_number(value, None)
            if margin is None or margin <= 0 or margin > 2:
                return False, None, f"'{self.MARGIN_SHEET}' 工作表 {self.MARGIN_CELL} 单元格的毛利率值无效 ({value})，请设置有效的毛利率（如 0.8411）"

            return True, margin, None

        except KeyError:
            return False, None, f"找不到工作表 '{self.MARGIN_SHEET}'"
        except Exception as e:
            return False, None, f"读取毛利率单元格时出错: {e}"

    def _get_margin(self):
        """获取 J14 单元格的毛利率值"""
        is_valid, margin, error_msg = self._check_margin_cell()
        if not is_valid:
            raise ValueError(error_msg)
        return margin

    def get_current_data(self):
        """获取当前数据（使用 formulas 读取实际计算值）"""
        self._load_model()
        try:
            solution = self._calculate()

            return {
                'E17': self._to_number(self._get_value(solution, '测算表', 'E17')),
                'E18': self._to_number(self._get_value(solution, '测算表', 'E18')),
                'E21': self._to_number(self._get_value(solution, '测算表', 'E21')),
                'E22': self._to_number(self._get_value(solution, '测算表', 'E22')),
                'E29': self._to_number(self._get_value(solution, '测算表', 'E29')),
                'E30': self._to_number(self._get_value(solution, '测算表', 'E30')),
                'E31': self._to_number(self._get_value(solution, '测算表', 'E31')),
                'G22': self._to_number(self._get_value(solution, '测算表', 'G22')),
                'B47': self._to_number(self._get_value(solution, '测算表', 'B47')),
                'G25': self._to_number(self._get_value(solution, '测算表', 'G25'), 1),
                'J12': self._to_number(self._get_value(solution, '销售成本', 'J12')),
                'B2': self._to_number(self._get_value(solution, '测算表', 'B2')),
            }
        finally:
            self._unload_model()

    def calculate_tax(self, income):
        """累进税率计算（个体工商户经营所得税率表）"""
        if income <= 30000:
            return income * 0.05
        elif income <= 90000:
            return income * 0.1 - 1500
        elif income <= 300000:
            return income * 0.2 - 10500
        elif income <= 500000:
            return income * 0.3 - 40500
        else:
            return income * 0.35 - 65500

    def reverse_calculate_income(self, tax):
        """根据税额反推应纳税所得额"""
        if tax <= 1500:
            return tax / 0.05
        elif tax <= 7500:
            return (tax + 1500) / 0.1
        elif tax <= 49500:
            return (tax + 10500) / 0.2
        elif tax <= 109500:
            return (tax + 40500) / 0.3
        else:
            return (tax + 65500) / 0.35

    def _get_G22_at_E18(self, E18):
        """设置 E18 并获取计算后的 G22 值"""
        inputs = {self._cell_key('测算表', 'E18'): E18}
        solution = self._calculate(inputs)
        return self._to_number(self._get_value(solution, '测算表', 'G22'))

    def _get_E31_at_G25(self, G25):
        """设置 G25 并获取计算后的 E31 值"""
        inputs = {self._cell_key('测算表', 'G25'): G25}
        solution = self._calculate(inputs)
        return self._to_number(self._get_value(solution, '测算表', 'E31'))

    def _get_B47_at_G25(self, G25):
        """设置 G25 并获取计算后的 B47 值"""
        inputs = {self._cell_key('测算表', 'G25'): G25}
        solution = self._calculate(inputs)
        return self._to_number(self._get_value(solution, '测算表', 'B47'))

    def find_E18_for_target_G22(self, target_G22=0, tolerance=0.009):
        """
        二分法查找 E18，使 G22 接近目标值
        返回 (E18, G22, is_in_range, boundary_info)
        """
        low, high = self.E18_MIN, self.E18_MAX

        # 先计算边界值
        G22_at_low = self._get_G22_at_E18(low)
        G22_at_high = self._get_G22_at_E18(high)

        # 确定 G22 随 E18 变化的方向
        # E18 增加 -> E21(税额) 增加 -> E22 增加 -> G22 减小
        # 所以 G22 随 E18 增加而减小

        # 检查目标是否在可达范围内
        min_G22 = min(G22_at_low, G22_at_high)
        max_G22 = max(G22_at_low, G22_at_high)

        if target_G22 < min_G22:
            # 目标太小，使用上界
            final_E18 = high if G22_at_high < G22_at_low else low
            final_G22 = self._get_G22_at_E18(final_E18)
            return final_E18, final_G22, False, {
                'reason': 'target_too_low',
                'target_G22': target_G22,
                'min_G22': min_G22,
            }
        if target_G22 > max_G22:
            # 目标太大，使用下界
            final_E18 = low if G22_at_low > G22_at_high else high
            final_G22 = self._get_G22_at_E18(final_E18)
            return final_E18, final_G22, False, {
                'reason': 'target_too_high',
                'target_G22': target_G22,
                'max_G22': max_G22,
            }

        # 二分法查找
        mid = (low + high) / 2
        for _ in range(50):  # 最多迭代50次
            if high - low < 0.01:  # E18 精度到 0.01 元（G22 敏感度约 0.1/元）
                break
            mid = (low + high) / 2
            G22 = self._get_G22_at_E18(mid)

            if abs(G22 - target_G22) < tolerance:
                return mid, G22, True, None

            # G22 随 E18 增加而减小
            if G22 > target_G22:
                low = mid
            else:
                high = mid

        final_G22 = self._get_G22_at_E18(mid)
        return mid, final_G22, abs(final_G22 - target_G22) < tolerance, None

    def find_G25_for_target_E31(self, target_E31=0, tolerance=0.009):
        """
        二分法查找 G25，使 E31 接近目标值
        返回 (G25, E31, B47, J12, is_in_range, boundary_info)
        """
        low, high = self.G25_MIN, self.G25_MAX

        def get_values_at_G25(g25):
            inputs = {self._cell_key('测算表', 'G25'): g25}
            solution = self._calculate(inputs)
            return (
                self._to_number(self._get_value(solution, '测算表', 'E31')),
                self._to_number(self._get_value(solution, '测算表', 'B47')),
                self._to_number(self._get_value(solution, '销售成本', 'J12')),
            )

        # 先计算边界值
        E31_at_low, B47_at_low, J12_at_low = get_values_at_G25(low)
        E31_at_high, B47_at_high, J12_at_high = get_values_at_G25(high)

        # E31 = E30 - E29, E30 = prev_profit + B47
        # G25 增加 -> 销售成本 J12 增加 -> B47 减小 -> E30 减小 -> E31 减小

        min_E31 = min(E31_at_low, E31_at_high)
        max_E31 = max(E31_at_low, E31_at_high)

        if target_E31 < min_E31:
            final_G25 = high if E31_at_high < E31_at_low else low
            E31, B47, J12 = get_values_at_G25(final_G25)
            return final_G25, E31, B47, J12, False, {
                'reason': 'target_too_low',
                'target_E31': target_E31,
                'min_E31': min_E31,
                'boundary_G25': final_G25,
            }
        if target_E31 > max_E31:
            final_G25 = low if E31_at_low > E31_at_high else high
            E31, B47, J12 = get_values_at_G25(final_G25)
            return final_G25, E31, B47, J12, False, {
                'reason': 'target_too_high',
                'target_E31': target_E31,
                'max_E31': max_E31,
                'boundary_G25': final_G25,
            }

        # 二分法查找
        mid = (low + high) / 2
        for _ in range(50):
            if high - low < 1e-9:
                break
            mid = (low + high) / 2
            E31, B47, J12 = get_values_at_G25(mid)

            if abs(E31 - target_E31) < tolerance:
                return mid, E31, B47, J12, True, None

            # E31 随 G25 增加而减小
            if E31 > target_E31:
                low = mid
            else:
                high = mid

        E31, B47, J12 = get_values_at_G25(mid)
        return mid, E31, B47, J12, abs(E31 - target_E31) < tolerance, None

    def calculate_combined_adjustment(self):
        """
        整合计算年利润和月毛利调整方案
        同时调整 E18 和 G25，使 G22 = 0 且 E31 = 0

        利用 formulas 公式计算，不在代码中硬编码公式
        """
        self._report_progress(0, "正在加载 Excel 模型...")
        self._load_model()
        try:
            self._report_progress(10, "正在读取当前数据...")
            # 获取当前数据（无输入修改）
            solution = self._calculate()

            original_E18 = self._to_number(self._get_value(solution, '测算表', 'E18'))
            original_G25 = self._to_number(self._get_value(solution, '测算表', 'G25'), 1)

            current = {
                'E17': self._to_number(self._get_value(solution, '测算表', 'E17')),
                'E18': original_E18,
                'E21': self._to_number(self._get_value(solution, '测算表', 'E21')),
                'E22': self._to_number(self._get_value(solution, '测算表', 'E22')),
                'G22': self._to_number(self._get_value(solution, '测算表', 'G22')),
                'G25': original_G25,
                'B47': self._to_number(self._get_value(solution, '测算表', 'B47')),
                'E29': self._to_number(self._get_value(solution, '测算表', 'E29')),
                'E30': self._to_number(self._get_value(solution, '测算表', 'E30')),
                'E31': self._to_number(self._get_value(solution, '测算表', 'E31')),
                'J12': self._to_number(self._get_value(solution, '销售成本', 'J12')),
            }

            # === 第一步: 查找 E18 使 G22 = 0 ===
            self._report_progress(20, "正在搜索最优 E18...")
            target_E18, verify_G22, e18_in_range, e18_boundary = self.find_E18_for_target_G22(
                target_G22=0, tolerance=0.009
            )

            # 检查 E18 是否在安全范围
            e18_safe, e18_msg = self._check_range(
                target_E18, self.E18_MIN, self.E18_MAX, 'E18'
            )

            # 设置新的 E18，计算新的 E29
            inputs_e18 = {self._cell_key('测算表', 'E18'): target_E18}
            solution_e18 = self._calculate(inputs_e18)
            new_E29 = self._to_number(self._get_value(solution_e18, '测算表', 'E29'))
            verify_E21 = self._to_number(self._get_value(solution_e18, '测算表', 'E21'))
            verify_E22 = self._to_number(self._get_value(solution_e18, '测算表', 'E22'))

            # === 第二步: 查找 G25 使 E31 = 0（在新 E18 基础上）===
            self._report_progress(50, "正在搜索最优 G25...")
            # 需要同时设置 E18 和 G25
            def get_values_at_G25_with_E18(g25):
                inputs = {
                    self._cell_key('测算表', 'E18'): target_E18,
                    self._cell_key('测算表', 'G25'): g25,
                }
                sol = self._calculate(inputs)
                return (
                    self._to_number(self._get_value(sol, '测算表', 'E31')),
                    self._to_number(self._get_value(sol, '测算表', 'B47')),
                    self._to_number(self._get_value(sol, '销售成本', 'J12')),
                )

            low, high = self.G25_MIN, self.G25_MAX

            E31_at_low, B47_at_low, J12_at_low = get_values_at_G25_with_E18(low)
            E31_at_high, B47_at_high, J12_at_high = get_values_at_G25_with_E18(high)

            min_E31 = min(E31_at_low, E31_at_high)
            max_E31 = max(E31_at_low, E31_at_high)
            target_E31 = 0
            tolerance = 0.009

            g25_in_range = True
            g25_boundary = None

            if target_E31 < min_E31:
                target_G25 = high if E31_at_high < E31_at_low else low
                verify_E31, verify_B47, verify_J12 = get_values_at_G25_with_E18(target_G25)
                g25_in_range = False
                g25_boundary = {
                    'reason': 'target_too_low',
                    'target_E31': target_E31,
                    'min_E31': min_E31,
                    'boundary_G25': target_G25,
                }
            elif target_E31 > max_E31:
                target_G25 = low if E31_at_low > E31_at_high else high
                verify_E31, verify_B47, verify_J12 = get_values_at_G25_with_E18(target_G25)
                g25_in_range = False
                g25_boundary = {
                    'reason': 'target_too_high',
                    'target_E31': target_E31,
                    'max_E31': max_E31,
                    'boundary_G25': target_G25,
                }
            else:
                # 二分法查找
                mid = (low + high) / 2
                for _ in range(50):
                    if high - low < 1e-9:
                        break
                    mid = (low + high) / 2
                    E31, B47, J12 = get_values_at_G25_with_E18(mid)

                    if abs(E31 - target_E31) < tolerance:
                        break

                    if E31 > target_E31:
                        low = mid
                    else:
                        high = mid

                target_G25 = mid
                verify_E31, verify_B47, verify_J12 = get_values_at_G25_with_E18(target_G25)
                g25_in_range = abs(verify_E31 - target_E31) < tolerance

            # 检查 G25 是否在安全范围
            g25_safe, g25_msg = self._check_range(
                target_G25, self.G25_MIN, self.G25_MAX, 'G25'
            )

            # 读取最终的 E30
            self._report_progress(90, "正在验证结果...")
            final_inputs = {
                self._cell_key('测算表', 'E18'): target_E18,
                self._cell_key('测算表', 'G25'): target_G25,
            }
            final_solution = self._calculate(final_inputs)
            verify_E30 = self._to_number(self._get_value(final_solution, '测算表', 'E30'))

            # 如果 target 值与 original 值非常接近，使用原始计算值以避免 formulas 库的计算差异
            e18_unchanged = abs(target_E18 - original_E18) < 1
            g25_unchanged = abs(target_G25 - original_G25) < 1e-5
            if e18_unchanged and g25_unchanged:
                # 当没有实际调整时，使用原始值
                verify_G22 = current['G22']
                verify_E31 = current['E31']
                verify_B47 = current['B47']
                verify_J12 = current['J12']
                new_E29 = current['E29']
                verify_E30 = current['E30']
                verify_E21 = current['E21']
                verify_E22 = current['E22']

            self._report_progress(100, "计算完成")
            # 构建结果
            result = {
                'current': current,
                'target': {
                    'E18': target_E18,
                    'G25': target_G25,
                    'B47': verify_B47,
                },
                'verify': {
                    'E21': verify_E21,
                    'E22': verify_E22,
                    'G22': verify_G22,
                    'B47': verify_B47,
                    'E29': new_E29,
                    'E30': verify_E30,
                    'E31': verify_E31,
                    'J12': verify_J12,
                },
                'in_range': e18_in_range and g25_in_range,
                'safety_check': {
                    'E18_safe': e18_safe,
                    'E18_msg': e18_msg,
                    'G25_safe': g25_safe,
                    'G25_msg': g25_msg,
                },
            }

            if not e18_in_range and e18_boundary:
                result['e18_boundary_info'] = e18_boundary
            if not g25_in_range and g25_boundary:
                result['boundary_info'] = g25_boundary

            return result

        finally:
            self._unload_model()

    def find_optimal_margin_fast(self, target_H11=0, target_F20=0, h11_tolerance=1.0, f20_tolerance=100):
        """
        快速搜索最优毛利率和B11（多目标帕累托优化）

        算法原理：
        1. F20 对 B11 是完美线性的（实验验证差分恒定）
        2. 沿着帕累托前沿搜索，找到最接近目标的平衡点
        3. 使用加权目标函数: error = w1*|H11| + w2*|F20|

        Args:
            target_H11: H11 目标值，默认 0
            target_F20: F20 目标值，默认 0
            h11_tolerance: H11 容差，默认 1.0
            f20_tolerance: F20 容差，默认 100

        Returns:
            dict: 包含 margin, B11, H11, F20, converged, iterations
        """
        calc_count = 0
        cache = {}

        def get_values(margin, b11):
            """获取 H11 和 F20 值（带缓存）"""
            nonlocal calc_count
            cache_key = (round(margin, 5), round(b11, 0))
            if cache_key in cache:
                return cache[cache_key]

            calc_count += 1
            inputs = {
                self._cell_key(self.MARGIN_SHEET, self.MARGIN_CELL): margin,
                self._cell_key('产品成本', 'B11'): b11,
            }
            solution = self._calculate(inputs)
            h11 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'H11'))
            f20 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'F20'))
            cache[cache_key] = (h11, f20)
            return h11, f20

        def find_b11_for_target_f20(margin, target_f20):
            """
            对于给定的 margin，找到使 F20=target_f20 的 B11
            利用 F20 对 B11 的线性特性
            """
            b11_1, b11_2 = 0, 200000
            _, f20_1 = get_values(margin, b11_1)
            _, f20_2 = get_values(margin, b11_2)

            slope = (f20_2 - f20_1) / (b11_2 - b11_1)

            if abs(slope) < 1e-10:
                return b11_1, slope

            target_b11 = b11_1 + (target_f20 - f20_1) / slope
            target_b11 = max(self.B11_MIN, min(self.B11_MAX, target_b11))

            return target_b11, slope

        def weighted_error(h11, f20, h11_weight=100.0, f20_weight=1.0):
            """计算加权误差（H11 优先，因为 H11 容差 ±10 比 F20 容差 ±40000 更严格）"""
            # H11 安全范围 ±10，F20 安全范围 ±40000
            h11_norm = abs(h11 - target_H11) / 10.0  # 归一化到容差
            f20_norm = abs(f20 - target_F20) / 40000.0  # 归一化到容差
            return h11_weight * h11_norm + f20_weight * f20_norm

        # 步骤1: 粗略扫描，确定搜索方向和最佳区域
        self._report_progress(20, "粗略扫描搜索空间...")
        # 更密集的网格，确保覆盖更多可能的最优区域
        margin_samples = [0.70, 0.72, 0.74, 0.76, 0.78, 0.80, 0.82, 0.84, 0.86, 0.88, 0.90]
        b11_samples = [0, 50000, 100000, 150000, 200000, 250000, 300000, 350000, 400000, 450000, 500000]

        best_result = None
        best_error = float('inf')
        all_samples = []  # 记录所有采样点，用于后续分析

        for margin in margin_samples:
            for b11 in b11_samples:
                h11, f20 = get_values(margin, b11)
                error = weighted_error(h11, f20)
                all_samples.append({
                    'margin': margin, 'B11': b11, 'H11': h11, 'F20': f20, 'error': error
                })

                if error < best_error:
                    best_error = error
                    best_result = {
                        'margin': margin,
                        'B11': b11,
                        'H11': h11,
                        'F20': f20,
                    }

        # 步骤2: 在最佳区域附近精细搜索
        self._report_progress(50, "精细搜索最优解...")
        if best_result:
            center_margin = best_result['margin']
            center_b11 = best_result['B11']

            # 精细网格
            margin_range = [center_margin + d * 0.01 for d in range(-5, 6)]
            margin_range = [m for m in margin_range if self.MARGIN_MIN <= m <= self.MARGIN_MAX]

            b11_range = [center_b11 + d * 20000 for d in range(-5, 6)]
            b11_range = [b for b in b11_range if self.B11_MIN <= b <= self.B11_MAX]

            for margin in margin_range:
                for b11 in b11_range:
                    h11, f20 = get_values(margin, b11)
                    error = weighted_error(h11, f20)

                    if error < best_error:
                        best_error = error
                        best_result = {
                            'margin': margin,
                            'B11': b11,
                            'H11': h11,
                            'F20': f20,
                        }

        # 步骤3: 梯度下降微调
        self._report_progress(70, "梯度下降微调...")
        if best_result:
            margin = best_result['margin']
            b11 = best_result['B11']

            for iteration in range(20):
                # 计算数值梯度
                delta_m = 0.001
                delta_b = 1000

                h11_0, f20_0 = get_values(margin, b11)
                error_0 = weighted_error(h11_0, f20_0)

                # margin 方向梯度
                if margin + delta_m <= self.MARGIN_MAX:
                    h11_m, f20_m = get_values(margin + delta_m, b11)
                    grad_m = (weighted_error(h11_m, f20_m) - error_0) / delta_m
                else:
                    grad_m = 0

                # B11 方向梯度
                if b11 + delta_b <= self.B11_MAX:
                    h11_b, f20_b = get_values(margin, b11 + delta_b)
                    grad_b = (weighted_error(h11_b, f20_b) - error_0) / delta_b
                else:
                    grad_b = 0

                # 梯度下降步进
                step_m = 0.005
                step_b = 10000

                new_margin = margin - step_m * (1 if grad_m > 0 else -1 if grad_m < 0 else 0)
                new_b11 = b11 - step_b * (1 if grad_b > 0 else -1 if grad_b < 0 else 0)

                new_margin = max(self.MARGIN_MIN, min(self.MARGIN_MAX, new_margin))
                new_b11 = max(self.B11_MIN, min(self.B11_MAX, new_b11))

                h11_new, f20_new = get_values(new_margin, new_b11)
                error_new = weighted_error(h11_new, f20_new)

                if error_new < best_error:
                    best_error = error_new
                    margin = new_margin
                    b11 = new_b11
                    best_result = {
                        'margin': margin,
                        'B11': b11,
                        'H11': h11_new,
                        'F20': f20_new,
                    }
                else:
                    # 收敛
                    break

        # 判断是否收敛
        h11_ok = abs(best_result['H11'] - target_H11) < h11_tolerance
        f20_ok = abs(best_result['F20'] - target_F20) < f20_tolerance
        converged = h11_ok and f20_ok

        best_result['converged'] = converged
        best_result['iterations'] = calc_count
        best_result['h11_converged'] = h11_ok
        best_result['f20_converged'] = f20_ok

        return best_result

    def find_optimal_margin_v2(self, h11_range, f20_range, margin_range):
        """
        优化版搜索算法：利用 F20-B11 线性关系 + 二分法

        原理：
        - 对于任意固定的 margin 值，F20 与 B11 是完美线性关系
        - 只需 2 次计算即可确定直线方程，直接求解使 F20 落在目标范围内的 B11
        - 对 margin 使用二分法搜索，找到使 H11 落在目标范围内的值

        Args:
            h11_range: (min, max) H11 目标范围
            f20_range: (min, max) F20 目标范围
            margin_range: (min, max) 毛利率搜索范围

        Returns:
            dict: 包含 margin, B11, H11, F20, converged, iterations
        """
        h11_min, h11_max = h11_range
        f20_min, f20_max = f20_range
        margin_min, margin_max = margin_range

        calc_count = 0
        best_result = None
        best_error = float('inf')

        def get_values(margin, b11):
            """获取 H11 和 F20 值"""
            nonlocal calc_count
            calc_count += 1
            inputs = {
                self._cell_key(self.MARGIN_SHEET, self.MARGIN_CELL): margin,
                self._cell_key('产品成本', 'B11'): b11,
            }
            solution = self._calculate(inputs)
            h11 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'H11'))
            f20 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'F20'))
            return h11, f20

        def find_b11_for_target_f20(margin, target_f20):
            """利用线性关系计算使 F20=target 的 B11"""
            # 采样两点确定直线: F20 = k * B11 + b
            b11_sample_1, b11_sample_2 = 0, 100000
            _, f20_1 = get_values(margin, b11_sample_1)
            _, f20_2 = get_values(margin, b11_sample_2)

            # 计算斜率
            k = (f20_2 - f20_1) / (b11_sample_2 - b11_sample_1)
            b = f20_1  # 截距 (当 B11=0 时)

            if abs(k) < 1e-10:
                # 斜率太小，F20 几乎不随 B11 变化
                return b11_sample_1

            # 求解 B11 = (target_f20 - b) / k
            target_b11 = (target_f20 - b) / k
            # 约束到安全范围
            target_b11 = max(self.B11_MIN, min(self.B11_MAX, target_b11))
            return target_b11

        # 二分法搜索 margin
        self._report_progress(20, "二分法搜索最优毛利率...")
        low, high = margin_min, margin_max
        target_f20 = (f20_min + f20_max) / 2  # F20 目标中点

        for iteration in range(25):  # 最多 25 次迭代
            mid = (low + high) / 2

            # 利用线性关系直接计算 B11
            b11 = find_b11_for_target_f20(mid, target_f20)

            # 验证结果
            h11, f20 = get_values(mid, b11)
            error = abs(h11) / 10.0 + abs(f20) / 40000.0

            # 更新最优解
            if error < best_error:
                best_error = error
                best_result = {
                    'margin': mid,
                    'B11': b11,
                    'H11': h11,
                    'F20': f20,
                }

            # 检查是否满足约束
            h11_ok = h11_min <= h11 <= h11_max
            f20_ok = f20_min <= f20 <= f20_max

            if h11_ok and f20_ok:
                # 找到满足约束的解
                best_result['converged'] = True
                best_result['iterations'] = calc_count
                return best_result

            # 调整搜索范围
            # H11 通常随 margin 增加而增加
            if h11 < (h11_min + h11_max) / 2:
                low = mid
            else:
                high = mid

            # 检查收敛
            if high - low < 0.0001:
                break

            self._report_progress(20 + int(iteration * 2.5), f"搜索中... (迭代 {iteration + 1})")

        # 未找到精确解，返回最优近似解
        if best_result is None:
            best_result = {
                'margin': (margin_min + margin_max) / 2,
                'B11': 0,
                'H11': 0,
                'F20': 0,
            }

        best_result['converged'] = False
        best_result['iterations'] = calc_count
        return best_result

    def find_alternative_solutions(self, optimal_result, target_H11=0, target_F20=0, num_alternatives=4):
        """
        生成帕累托前沿上的备选方案

        提供不同权衡的方案：
        - H11 优先方案
        - F20 优先方案
        - 平衡方案

        Args:
            optimal_result: find_optimal_margin_fast 的返回结果
            target_H11: H11 目标值
            target_F20: F20 目标值
            num_alternatives: 备选方案数量

        Returns:
            list: 备选方案列表
        """
        alternatives = []

        def get_values(margin, b11):
            inputs = {
                self._cell_key(self.MARGIN_SHEET, self.MARGIN_CELL): margin,
                self._cell_key('产品成本', 'B11'): b11,
            }
            solution = self._calculate(inputs)
            h11 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'H11'))
            f20 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'F20'))
            return h11, f20

        # 策略1: H11 优先（固定较小的 B11，调整 margin 使 H11 最小）
        self._report_progress(86, "生成 H11 优先方案...")
        best_h11_sol = None
        best_h11_error = float('inf')

        for margin in [0.70, 0.72, 0.74, 0.76, 0.78, 0.80]:
            for b11 in [0, 50000, 100000]:
                h11, f20 = get_values(margin, b11)
                h11_error = abs(h11 - target_H11)
                if h11_error < best_h11_error:
                    best_h11_error = h11_error
                    best_h11_sol = {'margin': margin, 'B11': b11, 'H11': h11, 'F20': f20}

        if best_h11_sol:
            best_h11_sol['label'] = 'H11优先'
            alternatives.append(best_h11_sol)

        # 策略2: F20 优先（使用线性关系直接计算使 F20=0 的 B11）
        self._report_progress(88, "生成 F20 优先方案...")
        for margin in [0.75, 0.80, 0.85]:
            # 采样两点确定线性关系
            _, f20_0 = get_values(margin, 0)
            _, f20_200k = get_values(margin, 200000)
            slope = (f20_200k - f20_0) / 200000

            if abs(slope) > 1e-10:
                target_b11 = -f20_0 / slope
                target_b11 = max(self.B11_MIN, min(self.B11_MAX, target_b11))
                h11, f20 = get_values(margin, target_b11)
                alternatives.append({
                    'margin': margin,
                    'B11': target_b11,
                    'H11': h11,
                    'F20': f20,
                    'label': f'F20优先 (m={margin:.2f})'
                })

            if len(alternatives) >= num_alternatives:
                break

        return alternatives[:num_alternatives]

    def calculate_inventory_margin_adjustment(self, max_solutions=5):
        """
        计算库存毛利率调整方案（优化版：利用F20线性特性快速搜索）
        目标: 使 H11 = 0.00, F20 = 0.00
        工作表: 生产成本月结表、产品成本
        调整变量: 毛利率 (J14单元格), B11 (产品成本中的加工费)

        优化原理：
        - F20 对 B11 是完美线性关系，可直接计算
        - 用二分法搜索 margin，每次迭代约 3 次计算
        - 总计算量从 ~600 次降低到 ~50 次

        Args:
            max_solutions: 最多返回的候选方案数量，默认 5

        Returns:
            dict: 包含 current（当前值）、solutions（方案列表）、stats（搜索统计）
        """
        self._report_progress(0, "正在加载 Excel 模型...")
        self._load_model()
        try:
            self._report_progress(5, "正在检查毛利率单元格...")
            # 检查 J14 单元格是否有效
            is_valid, current_margin, error_msg = self._check_margin_cell()
            if not is_valid:
                return {
                    'error': error_msg,
                    'current': {},
                    'solutions': [],
                }

            self._report_progress(10, "正在读取当前数据...")
            # 获取当前值
            solution = self._calculate()
            current_H11 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'H11'))
            current_F20 = self._to_number(self._get_value(solution, self.MARGIN_SHEET, 'F20'))
            current_B11 = self._to_number(self._get_value(solution, '产品成本', 'B11'))

            # 添加"保持当前值"作为候选方案
            current_solution = {
                'margin': current_margin,
                'B11': current_B11,
                'H11': current_H11,
                'F20': current_F20,
                'label': '当前值'
            }

            # 使用优化的快速搜索算法
            self._report_progress(15, "正在快速搜索最优解...")
            optimal_result = self.find_optimal_margin_fast(
                target_H11=0, target_F20=0,
                h11_tolerance=1.0,
                f20_tolerance=100
            )

            # 构建解列表
            all_solutions = [current_solution]

            # 添加最优解
            if optimal_result.get('converged'):
                optimal_result['label'] = '最优解 ✓'
            else:
                boundary = optimal_result.get('boundary', '')
                if boundary == 'target_too_low':
                    optimal_result['label'] = '边界解 (H11下限)'
                elif boundary == 'target_too_high':
                    optimal_result['label'] = '边界解 (H11上限)'
                else:
                    optimal_result['label'] = '近似解'

            all_solutions.append(optimal_result)

            # 添加备选方案
            self._report_progress(85, "正在生成备选方案...")
            alternatives = self.find_alternative_solutions(
                optimal_result,
                target_H11=0, target_F20=0,
                num_alternatives=min(3, max_solutions - 2)
            )
            for i, alt in enumerate(alternatives):
                alt['label'] = f'备选方案 {i + 1}'
                all_solutions.append(alt)

            # 为每个解添加安全检查
            self._report_progress(95, "正在验证结果...")
            for sol in all_solutions:
                margin_safe, _ = self._check_range(sol['margin'], self.MARGIN_MIN, self.MARGIN_MAX, '毛利率')
                b11_safe, _ = self._check_range(sol['B11'], self.B11_MIN, self.B11_MAX, 'B11')
                h11_safe, _ = self._check_range(sol['H11'], self.H11_MIN, self.H11_MAX, 'H11')
                f20_safe, _ = self._check_range(sol['F20'], self.F20_MIN, self.F20_MAX, 'F20')

                sol['h11_ok'] = h11_safe
                sol['f20_ok'] = f20_safe
                sol['all_ok'] = margin_safe and b11_safe and h11_safe and f20_safe

            self._report_progress(100, "计算完成")
            return {
                'current': {
                    'margin': current_margin,
                    'H11': current_H11,
                    'F20': current_F20,
                    'B11': current_B11,
                },
                'solutions': all_solutions,
                'stats': {
                    'iterations': optimal_result.get('iterations', 0),
                    'converged': optimal_result.get('converged', False),
                },
            }

        except KeyError as e:
            return {
                'error': f'找不到工作表: {e}',
                'current': {},
                'solutions': [],
            }
        finally:
            self._unload_model()

    def _save_to_file(self, values):
        """用 openpyxl 保存值到文件

        Args:
            values: dict of {sheet_name: {cell: value}}
        """
        wb = openpyxl.load_workbook(self.temp_file_path)
        for sheet_name, cells in values.items():
            ws = wb[sheet_name]
            for cell, value in cells.items():
                ws[cell] = value
        wb.save(self.temp_file_path)
        wb.close()
