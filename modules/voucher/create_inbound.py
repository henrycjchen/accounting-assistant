import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .config import FLOAT_UNITS
from .helpers import set_wrap_border, random_range, random_pick


def create_inbound(workbook, file_path, outbound):
    """创建入库凭证"""
    source_wb = load_workbook(file_path, data_only=True)

    # 查找销售成本表
    if '销售成本' not in source_wb.sheetnames:
        raise Exception('未找到销售成本表')

    sheet = source_wb['销售成本']
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    result = wash_data(data)
    valid_data = result['valid_data']

    valid_data_formatted = format_data(valid_data, outbound)

    action(valid_data_formatted, workbook)

    return valid_data_formatted


def action(valid_data, workbook):
    """生成工作表"""
    worksheet = workbook.create_sheet('入库凭证')
    worksheet.print_options.horizontalCentered = True

    # 设置列宽
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 18.17
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 4.33
    worksheet.column_dimensions['E'].width = 20

    row = 1

    for index, items in enumerate(valid_data):
        # 标题行
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.value = '入  库  凭  证'
        cell.font = Font(bold=True, size=22)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        # Apply bottom border to all cells in merged range
        double_bottom = Border(bottom=Side(style='double'))
        for col in range(1, 6):
            worksheet.cell(row=row, column=col).border = double_bottom
        worksheet.row_dimensions[row].height = 38

        row += 1
        # 领取人和日期
        worksheet.merge_cells(f'A{row}:B{row}')
        cell = worksheet[f'A{row}']
        cell.value = '领取人：生产车间'
        cell.alignment = Alignment(vertical='center', wrap_text=True)

        date_cell = worksheet[f'C{row}']
        date_obj = datetime.fromtimestamp(items[0]['date'])
        date_cell.value = date_obj.strftime('%Y年%m月%d日')
        date_cell.alignment = Alignment(vertical='center')
        worksheet.row_dimensions[row].height = 30

        row += 1
        # 表头
        headers = ['用途', '品名', '规格', '单位', '数量']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            set_wrap_border(cell)
        worksheet.row_dimensions[row].height = 20

        # 数据行
        for idx, product in enumerate(items):
            row += 1
            cell_a = worksheet.cell(row=row, column=1)
            cell_a.value = '生产' if idx == 0 else ''
            set_wrap_border(cell_a)

            cell_b = worksheet.cell(row=row, column=2)
            cell_b.value = product['product']
            set_wrap_border(cell_b)

            cell_c = worksheet.cell(row=row, column=3)
            cell_c.value = ''
            set_wrap_border(cell_c)

            cell_d = worksheet.cell(row=row, column=4)
            cell_d.value = product['unit']
            set_wrap_border(cell_d)

            cell_e = worksheet.cell(row=row, column=5)
            if product['unit'] in FLOAT_UNITS:
                cell_e.value = round(product['count'], 3)
            else:
                cell_e.value = int(product['count'])
            set_wrap_border(cell_e)

            worksheet.row_dimensions[row].height = 20

        # 填充空行
        for _ in range(len(items), 7):
            row += 1
            for col in range(1, 6):
                set_wrap_border(worksheet.cell(row=row, column=col))
            worksheet.row_dimensions[row].height = 20

        # 合计行
        row += 1
        worksheet.merge_cells(f'A{row}:D{row}')
        cell = worksheet[f'A{row}']
        cell.value = f"合{' ' * 20}计"
        # Apply border to all cells in merged range (A-D)
        for col in range(1, 5):
            set_wrap_border(worksheet.cell(row=row, column=col))
        set_wrap_border(worksheet[f'E{row}'])
        worksheet.row_dimensions[row].height = 20

        # 保管人
        row += 1
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.alignment = Alignment(vertical='center', horizontal='right')
        cell.value = f"保管人：陈{' ' * 20}"
        worksheet.row_dimensions[row].height = 24

        if index % 2 == 0:
            row += 11
        else:
            row += 3

    # 设置右侧列宽
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 18.17
    worksheet.column_dimensions['J'].width = 20
    worksheet.column_dimensions['K'].width = 4.33
    worksheet.column_dimensions['L'].width = 20


def wash_data(data):
    """清洗数据"""
    count_target = find_target(data, '本期生产')
    product_target = find_target(data, '品名')

    if not count_target or not product_target:
        raise Exception('销售成本表未找到[本期生产]或[品名]')

    slim_data = []
    for item in data[count_target[0] + 2:]:
        if not item or len(item) == 0:
            continue
        product_val = item[product_target[1]] if len(item) > product_target[1] else None
        if not product_val:
            continue
        if re.search(r'合\s*计', str(product_val)):
            continue

        product_str = str(product_val).strip()
        parts = re.split(r'[(（]', product_str)
        product = parts[0].strip() if parts else ''
        unit = parts[1].strip() if len(parts) > 1 else ''
        unit = unit.rstrip(')）')

        count_val = item[count_target[1]] if len(item) > count_target[1] else 0
        try:
            count = float(count_val) if count_val else 0
        except (ValueError, TypeError):
            count = 0

        if count:
            slim_data.append({
                'product': product,
                'unit': unit,
                'count': count
            })

    return {'valid_data': slim_data}


def find_target(data, target):
    """查找目标单元格位置"""
    for i, row in enumerate(data):
        if not row:
            continue
        for j, cell in enumerate(row):
            if not cell:
                continue
            if target in str(cell).replace(' ', ''):
                return [i, j]
    return None


def merge_by_date(data):
    """按日期合并"""
    result = {}
    for items in data:
        for item in items:
            date_key = item['date']
            if date_key in result:
                result[date_key].append(item)
            else:
                result[date_key] = [item]
    return list(result.values())


def merge_by_product(data):
    """按产品合并"""
    result = []
    for items in data:
        product_map = {}
        for item in items:
            key = f"{item['product']}_{item['unit']}"
            if key in product_map:
                product_map[key]['count'] += item['count']
            else:
                product_map[key] = item.copy()
        result.append(list(product_map.values()))
    return result


def format_data(slim_data, outbound):
    """格式化数据"""
    merged_outbound = merge_by_product(merge_by_date(outbound))

    outbound_time_splitted = split_by_outbound_time(slim_data, merged_outbound)
    count_splitted = split_by_count(outbound_time_splitted)

    return count_splitted


def split_by_outbound_time(slim_data, outbound):
    """按出库时间拆分"""
    result = []

    # 过滤空列表
    outbound = [x for x in outbound if x]
    if not outbound:
        # 如果没有出库数据，返回所有入库数据作为一个批次
        return [slim_data] if slim_data else []

    inbound_count = random_range(5, 10)
    is_too_few = inbound_count > len(outbound)

    inbound_map = {}
    for item in slim_data:
        key = f"{item['product']}_{item['unit']}"
        inbound_map[key] = item.copy()

    first_date = datetime.fromtimestamp(outbound[0][0]['date'])
    pre_unix = int(first_date.replace(day=14, hour=23, minute=59, second=59).timestamp())

    for i in range(inbound_count):
        outbound_items = outbound[min(i, len(outbound) - 1)]

        # 更新日期
        pre_date = datetime.fromtimestamp(pre_unix)
        pre_date = pre_date.replace(hour=23, minute=59, second=59)
        pre_unix = int(pre_date.timestamp())

        two_days_later = pre_date + timedelta(days=2)
        outbound_date = datetime.fromtimestamp(outbound_items[0]['date']).replace(
            hour=0, minute=0, second=0
        )

        new_date = datetime.fromtimestamp(random_range(pre_unix, int(two_days_later.timestamp())))
        if new_date > outbound_date:
            new_date = outbound_date

        pre_unix = int(new_date.replace(hour=0, minute=0, second=0).timestamp())

        if i == inbound_count - 1:
            # 最后一批，使用所有剩余
            inbound = []
            for key, item in inbound_map.items():
                if item['count'] > 0:
                    inbound.append({
                        'date': pre_unix,
                        'product': item['product'],
                        'unit': item['unit'],
                        'count': item['count']
                    })
            result.append(inbound)
        else:
            inbound = []
            for item in outbound_items:
                if is_too_few and random_range(0, 1) < 0.5:
                    continue

                key = f"{item['product']}_{item['unit']}"
                if key not in inbound_map or inbound_map[key]['count'] <= 0:
                    continue

                if inbound_map[key]['count'] <= item['count'] and not is_too_few:
                    product_count = inbound_map[key]['count']
                else:
                    base_count = max(item['count'], inbound_map[key]['count'] / (inbound_count - i))
                    multiplier = 0.25 if is_too_few else 1
                    multiplier_max = 1.25 if is_too_few else 2
                    product_count = min(
                        random_range(
                            base_count * multiplier,
                            base_count * multiplier_max,
                            item['unit'] not in FLOAT_UNITS
                        ),
                        inbound_map[key]['count']
                    )

                inbound_map[key]['count'] -= product_count

                inbound.append({
                    'date': pre_unix,
                    'product': item['product'],
                    'unit': item['unit'],
                    'count': product_count
                })

            # 补充随机产品
            left_count = random_range(
                (7 - (len(inbound) % 7)) * 0.7,
                7 - (len(inbound) % 7)
            )

            if left_count > 0:
                outbound_products = [item['product'] for item in outbound_items]
                difference = [
                    item for key, item in inbound_map.items()
                    if item['count'] > 0 and item['product'] not in outbound_products
                ]
                random_products = random_pick(difference, int(left_count))

                for j, random_product in enumerate(random_products):
                    product_count = max(random_product['count'] / (inbound_count - i), 1)
                    random_count = min(
                        random_range(
                            product_count,
                            product_count * 2,
                            random_product['unit'] not in FLOAT_UNITS
                        ),
                        random_product['count']
                    )

                    inbound.append({
                        'date': pre_unix,
                        'product': random_product['product'],
                        'unit': random_product['unit'],
                        'count': random_count
                    })

                    random_product['count'] -= random_count

            result.append(inbound)

    return result


def split_by_count(data):
    """按数量拆分（每7条一组）"""
    result = []
    for items in data:
        count = (len(items) + 6) // 7
        for i in range(count):
            result.append(items[i * 7:(i + 1) * 7])
    return result
