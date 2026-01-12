import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

from .config import FLOAT_UNITS
from .helpers import set_wrap_border, random_range, random_pick


def create_issuing(workbook, file_path, inbound):
    """创建领料单"""
    source_wb = load_workbook(file_path, data_only=True)

    # 查找材料表
    if '材料' not in source_wb.sheetnames:
        raise Exception('未找到材料表')

    sheet = source_wb['材料']
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    result = wash_data(data)
    valid_data = result['valid_data']

    valid_data_formatted = format_data(valid_data, inbound)

    action(valid_data_formatted, workbook)

    return valid_data_formatted


def action(valid_data, workbook):
    """生成工作表"""
    worksheet = workbook.create_sheet('领料单')
    worksheet.print_options.horizontalCentered = True

    # 设置列宽
    worksheet.column_dimensions['A'].width = 27.63
    worksheet.column_dimensions['B'].width = 7.79
    worksheet.column_dimensions['C'].width = 17.79
    worksheet.column_dimensions['D'].width = 13.33
    worksheet.column_dimensions['E'].width = 13.33

    row = 1

    for index, items in enumerate(valid_data):
        # 标题行
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.value = '领  料  单'
        cell.font = Font(bold=True, size=22)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        # Apply bottom border to all cells in merged range
        double_bottom = Border(bottom=Side(style='double'))
        for col in range(1, 6):
            worksheet.cell(row=row, column=col).border = double_bottom
        worksheet.row_dimensions[row].height = 38

        row += 1
        # 日期
        worksheet.merge_cells(f'A{row}:E{row}')
        date_cell = worksheet[f'A{row}']
        date_obj = datetime.fromtimestamp(items[0]['date'])
        date_cell.value = date_obj.strftime('%Y年%m月%d日')
        date_cell.alignment = Alignment(vertical='center', horizontal='center')
        worksheet.row_dimensions[row].height = 30

        row += 1
        # 用料部门
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.value = '用料部门：生产车间                                 用途：生产'
        # Apply border to all cells in merged range (A-E)
        for col in range(1, 6):
            set_wrap_border(worksheet.cell(row=row, column=col))
        worksheet.row_dimensions[row].height = 20

        row += 1
        # 表头
        headers = ['材料名称及规格', '单位', '数量', '页次', '备注']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            set_wrap_border(cell)
        worksheet.row_dimensions[row].height = 20

        # 数据行
        for product in items:
            row += 1
            cell_a = worksheet.cell(row=row, column=1)
            cell_a.value = product['product']
            set_wrap_border(cell_a)

            cell_b = worksheet.cell(row=row, column=2)
            cell_b.value = product['unit']
            set_wrap_border(cell_b)

            cell_c = worksheet.cell(row=row, column=3)
            if product['unit'] in FLOAT_UNITS:
                cell_c.value = round(product['count'], 3)
            else:
                cell_c.value = int(product['count'])
            set_wrap_border(cell_c)

            cell_d = worksheet.cell(row=row, column=4)
            cell_d.value = ''
            set_wrap_border(cell_d)

            cell_e = worksheet.cell(row=row, column=5)
            cell_e.value = ''
            set_wrap_border(cell_e)

            worksheet.row_dimensions[row].height = 20

        # 填充空行
        for _ in range(len(items), 7):
            row += 1
            for col in range(1, 6):
                set_wrap_border(worksheet.cell(row=row, column=col))
            worksheet.row_dimensions[row].height = 20

        # 记账人
        row += 1
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.alignment = Alignment(vertical='center', horizontal='right')
        cell.value = f"记账：陈{' ' * 20}"
        worksheet.row_dimensions[row].height = 24

        if index % 2 == 0:
            row += 11
        else:
            row += 3

    # 设置右侧列宽
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 18.17
    worksheet.column_dimensions['J'].width = 20
    worksheet.column_dimensions['K'].width = 4.33
    worksheet.column_dimensions['L'].width = 20


def wash_data(data):
    """清洗数据"""
    count_target = find_target(data, '本月发出数')
    product_target = find_target(data, '品名')

    if not count_target or not product_target:
        raise Exception('材料表未找到[本月发出数]或[品名]')

    slim_data = []
    for item in data[count_target[0] + 2:]:
        if not item or len(item) == 0:
            continue
        product_val = item[product_target[1]] if len(item) > product_target[1] else None
        if not product_val:
            continue
        if re.search(r'合\s*计', str(product_val)):
            continue

        product = str(product_val).strip()
        unit_val = item[product_target[1] + 1] if len(item) > product_target[1] + 1 else ''
        unit = str(unit_val).strip() if unit_val else ''

        count_val = item[count_target[1]] if len(item) > count_target[1] else 0
        try:
            count = float(count_val) if count_val else 0
        except (ValueError, TypeError):
            count = 0

        if count:
            slim_data.append({
                'product': product,
                'unit': unit,
                'count': count
            })

    return {'valid_data': slim_data}


def find_target(data, target):
    """查找目标单元格位置"""
    for i, row in enumerate(data):
        if not row:
            continue
        for j, cell in enumerate(row):
            if not cell:
                continue
            if target in str(cell).replace(' ', ''):
                return [i, j]
    return None


def merge_by_date(data):
    """按日期合并"""
    result = {}
    for items in data:
        for item in items:
            date_key = item['date']
            if date_key in result:
                result[date_key].append(item)
            else:
                result[date_key] = [item]
    return list(result.values())


def format_data(slim_data, inbound):
    """格式化数据"""
    merged_inbound = merge_by_date(inbound)

    outbound_time_splitted = split_by_inbound_time(slim_data, merged_inbound)
    count_splitted = split_by_count(outbound_time_splitted)

    return count_splitted


def split_by_inbound_time(slim_data, inbound):
    """按入库时间拆分"""
    result = []

    # 过滤空列表
    inbound = [x for x in inbound if x]
    if not inbound:
        return [slim_data] if slim_data else []

    issuing_count = min(len(inbound), random_range(5, 8))

    issuing_map = {}
    for item in slim_data:
        key = f"{item['product']}_{item['unit']}"
        issuing_map[key] = item.copy()

    first_date = datetime.fromtimestamp(inbound[0][0]['date'])
    pre_unix = int(first_date.replace(day=9, hour=23, minute=59, second=59).timestamp())

    for i in range(issuing_count):
        # 更新日期
        pre_date = datetime.fromtimestamp(pre_unix)
        pre_date = pre_date.replace(hour=23, minute=59, second=59)
        pre_unix = int(pre_date.timestamp())

        two_days_later = pre_date + timedelta(days=2)
        inbound_date = datetime.fromtimestamp(inbound[i][0]['date']).replace(
            hour=0, minute=0, second=0
        )

        new_date = datetime.fromtimestamp(random_range(pre_unix, int(two_days_later.timestamp())))
        if new_date > inbound_date:
            new_date = inbound_date

        pre_unix = int(new_date.replace(hour=0, minute=0, second=0).timestamp())

        if i == issuing_count - 1:
            # 最后一批，使用所有剩余
            issuing = []
            for key, item in issuing_map.items():
                if item['count'] > 0:
                    issuing.append({
                        'date': pre_unix,
                        'product': item['product'],
                        'unit': item['unit'],
                        'count': item['count']
                    })
            result.append(issuing)
        else:
            products = [item for key, item in issuing_map.items() if item['count'] > 0]
            random_products = random_pick(
                products,
                random_range(max(1, int(len(slim_data) * 0.5)), len(slim_data))
            )

            issuing = []
            for item in random_products:
                product_count = min(
                    random_range(
                        item['count'] / (issuing_count - i),
                        (item['count'] / (issuing_count - i)) * 2,
                        item['unit'] not in FLOAT_UNITS
                    ),
                    item['count']
                )
                item['count'] -= product_count

                issuing.append({
                    'date': pre_unix,
                    'product': item['product'],
                    'unit': item['unit'],
                    'count': product_count
                })

            result.append(issuing)

    return result


def split_by_count(data):
    """按数量拆分（每7条一组）"""
    result = []
    for items in data:
        count = (len(items) + 6) // 7
        for i in range(count):
            result.append(items[i * 7:(i + 1) * 7])
    return result
