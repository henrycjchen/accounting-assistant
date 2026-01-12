from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import locale

from .config import FLOAT_UNITS
from .helpers import set_wrap_border
from .handle_outbound_data import handle_outbound_data


def create_outbound(workbook, file_path):
    """创建出库凭证"""
    result = handle_outbound_data(file_path)
    valid_data = result['valid_data']
    invalid_data = result['invalid_data']

    valid_data_formatted = format_data(valid_data)
    invalid_data_formatted = format_data(invalid_data)

    action(valid_data_formatted, invalid_data_formatted, workbook)

    return valid_data_formatted


def action(valid_data, invalid_data, workbook):
    """生成工作表"""
    worksheet = workbook.create_sheet('出库凭证')
    worksheet.print_options.horizontalCentered = True

    # 设置列宽
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 18.17
    worksheet.column_dimensions['C'].width = 20
    worksheet.column_dimensions['D'].width = 4.33
    worksheet.column_dimensions['E'].width = 20

    row = 1

    for index, items in enumerate(valid_data):
        # 标题行
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.value = '出  库  凭  证'
        cell.font = Font(bold=True, size=22)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        # Apply bottom border to all cells in merged range
        double_bottom = Border(bottom=Side(style='double'))
        for col in range(1, 6):
            worksheet.cell(row=row, column=col).border = double_bottom
        worksheet.row_dimensions[row].height = 38

        row += 1
        # 领取人和日期
        worksheet.merge_cells(f'A{row}:B{row}')
        cell = worksheet[f'A{row}']
        cell.value = f"领取人：{items[0]['buy_company']}"
        cell.alignment = Alignment(vertical='center', wrap_text=True)

        date_cell = worksheet[f'C{row}']
        date_obj = datetime.fromtimestamp(items[0]['date'])
        date_cell.value = date_obj.strftime('%Y年%m月%d日')
        date_cell.alignment = Alignment(vertical='center')
        worksheet.row_dimensions[row].height = 30

        row += 1
        # 表头
        headers = ['用途', '品名', '规格', '单位', '数量']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            set_wrap_border(cell)
        worksheet.row_dimensions[row].height = 20

        # 数据行
        for idx, product in enumerate(items):
            row += 1
            cell_a = worksheet.cell(row=row, column=1)
            cell_a.value = '销售' if idx == 0 else ''
            set_wrap_border(cell_a)

            cell_b = worksheet.cell(row=row, column=2)
            cell_b.value = product['product']
            set_wrap_border(cell_b)

            cell_c = worksheet.cell(row=row, column=3)
            cell_c.value = ''
            set_wrap_border(cell_c)

            cell_d = worksheet.cell(row=row, column=4)
            cell_d.value = product['unit']
            set_wrap_border(cell_d)

            cell_e = worksheet.cell(row=row, column=5)
            if product['unit'] in FLOAT_UNITS:
                cell_e.value = round(product['count'], 3)
            else:
                cell_e.value = product['count']
            set_wrap_border(cell_e)

            worksheet.row_dimensions[row].height = 20

        # 填充空行
        for _ in range(len(items), 7):
            row += 1
            for col in range(1, 6):
                set_wrap_border(worksheet.cell(row=row, column=col))
            worksheet.row_dimensions[row].height = 20

        # 合计行
        row += 1
        worksheet.merge_cells(f'A{row}:D{row}')
        cell = worksheet[f'A{row}']
        cell.value = f"合{' ' * 20}计"
        # Apply border to all cells in merged range (A-D)
        for col in range(1, 5):
            set_wrap_border(worksheet.cell(row=row, column=col))
        set_wrap_border(worksheet[f'E{row}'])
        worksheet.row_dimensions[row].height = 20

        # 保管人
        row += 1
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.alignment = Alignment(vertical='center', horizontal='right')
        cell.value = f"保管人：陈{' ' * 20}"
        worksheet.row_dimensions[row].height = 24

        if index % 2 == 0:
            row += 11
        else:
            row += 3

    # 入库凭证（退货）部分
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 18.17
    worksheet.column_dimensions['J'].width = 20
    worksheet.column_dimensions['K'].width = 4.33
    worksheet.column_dimensions['L'].width = 20

    row = 1
    for index, items in enumerate(invalid_data):
        # 标题行
        worksheet.merge_cells(f'H{row}:L{row}')
        cell = worksheet[f'H{row}']
        cell.value = '入  库  凭  证'
        cell.font = Font(bold=True, size=22)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        # Apply bottom border to all cells in merged range (columns H-L = 8-12)
        double_bottom = Border(bottom=Side(style='double'))
        for col in range(8, 13):
            worksheet.cell(row=row, column=col).border = double_bottom
        worksheet.row_dimensions[row].height = 38

        row += 1
        # 送货人和日期
        worksheet.merge_cells(f'H{row}:I{row}')
        cell = worksheet[f'H{row}']
        cell.value = f"送货人：{items[0]['buy_company']}"
        cell.alignment = Alignment(vertical='center', wrap_text=True)

        date_cell = worksheet[f'J{row}']
        date_obj = datetime.fromtimestamp(items[0]['date'])
        date_cell.value = date_obj.strftime('%Y年%m月%d日')
        date_cell.alignment = Alignment(vertical='center')
        worksheet.row_dimensions[row].height = 30

        row += 1
        # 表头
        cols = ['H', 'I', 'J', 'K', 'L']
        headers = ['用途', '品名', '规格', '单位', '数量']
        for col, header in zip(cols, headers):
            cell = worksheet[f'{col}{row}']
            cell.value = header
            set_wrap_border(cell)
        worksheet.row_dimensions[row].height = 20

        # 数据行
        for idx, product in enumerate(items):
            row += 1
            cell_h = worksheet[f'H{row}']
            cell_h.value = '退货入库' if idx == 0 else ''
            set_wrap_border(cell_h)

            cell_i = worksheet[f'I{row}']
            cell_i.value = product['product']
            set_wrap_border(cell_i)

            cell_j = worksheet[f'J{row}']
            cell_j.value = ''
            set_wrap_border(cell_j)

            cell_k = worksheet[f'K{row}']
            cell_k.value = product['unit']
            set_wrap_border(cell_k)

            cell_l = worksheet[f'L{row}']
            cell_l.value = -product['count']
            set_wrap_border(cell_l)

            worksheet.row_dimensions[row].height = 20

        # 填充空行
        for _ in range(len(items), 7):
            row += 1
            for col in cols:
                set_wrap_border(worksheet[f'{col}{row}'])
            worksheet.row_dimensions[row].height = 20

        # 合计行
        row += 1
        worksheet.merge_cells(f'H{row}:K{row}')
        cell = worksheet[f'H{row}']
        cell.value = f"合{' ' * 20}计"
        # Apply border to all cells in merged range (H-K = 8-11)
        for col in range(8, 12):
            set_wrap_border(worksheet.cell(row=row, column=col))
        set_wrap_border(worksheet[f'L{row}'])
        worksheet.row_dimensions[row].height = 20

        # 保管人
        row += 1
        worksheet.merge_cells(f'H{row}:L{row}')
        cell = worksheet[f'H{row}']
        cell.alignment = Alignment(vertical='center', horizontal='right')
        cell.value = f"保管人：陈{' ' * 20}"
        worksheet.row_dimensions[row].height = 24

        if index % 2 == 0:
            row += 11
        else:
            row += 3


def format_data(slim_data):
    """格式化数据"""
    company_splitted = merge_by_company(slim_data)
    date_splitted = split_by_date(company_splitted)
    count_merged = merge_counts(date_splitted)
    count_splitted = split_by_count(count_merged)
    date_sorted = sort_by_date(count_splitted)

    return date_sorted


def merge_by_company(data):
    """按公司合并"""
    result = {}
    for item in data:
        company = item['buy_company']
        if company in result:
            result[company].append(item)
        else:
            result[company] = [item]
    return list(result.values())


def split_by_date(data):
    """按日期拆分"""
    # 先按日期排序
    for items in data:
        items.sort(key=lambda x: x['date'])

    result = []
    for items in data:
        date_map = {}
        for item in items:
            date_key = item['date']
            if date_key in date_map:
                date_map[date_key].append(item)
            else:
                date_map[date_key] = [item]
        result.extend(list(date_map.values()))

    return result


def sort_by_date(data):
    """按日期排序"""
    # 过滤空列表
    non_empty = [x for x in data if x]
    return sorted(non_empty, key=lambda x: x[0]['date'])


def merge_counts(data):
    """合并数量"""
    result = []
    for items in data:
        # 按产品名称排序（中文）
        try:
            locale.setlocale(locale.LC_COLLATE, 'zh_CN.UTF-8')
            items.sort(key=lambda x: locale.strxfrm(x['product']))
        except:
            items.sort(key=lambda x: x['product'])

        product_map = {}
        for item in items:
            key = f"{item['product']}_{item['unit']}"
            if key in product_map:
                product_map[key]['count'] += item['count']
            else:
                product_map[key] = item.copy()
        result.append(list(product_map.values()))

    return result


def split_by_count(data):
    """按数量拆分（每7条一组）"""
    result = []
    for items in data:
        count = (len(items) + 6) // 7  # 向上取整
        for i in range(count):
            result.append(items[i * 7:(i + 1) * 7])
    return result
