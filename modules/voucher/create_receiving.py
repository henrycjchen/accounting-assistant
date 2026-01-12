import locale
from datetime import datetime, timedelta
from openpyxl.styles import Font, Alignment, Border, Side

from .config import FLOAT_UNITS
from .helpers import set_wrap_border, random_range
from .handle_inbound_data import handle_inbound_data


def create_receiving(workbook, file_path, issuing):
    """创建收料单"""
    result = handle_inbound_data(file_path)
    valid_data = result['valid_data']

    valid_data_formatted = format_data(valid_data, issuing)

    action(valid_data_formatted, workbook)

    return valid_data_formatted


def action(valid_data, workbook):
    """生成工作表"""
    worksheet = workbook.create_sheet('收料单')
    worksheet.print_options.horizontalCentered = True

    # 设置列宽
    worksheet.column_dimensions['A'].width = 36.75
    worksheet.column_dimensions['B'].width = 18.93
    worksheet.column_dimensions['C'].width = 8.14
    worksheet.column_dimensions['D'].width = 13.33
    worksheet.column_dimensions['E'].width = 8.33

    row = 1

    for index, items in enumerate(valid_data):
        # 标题行
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.value = '收  料  单'
        cell.font = Font(bold=True, size=22)
        cell.alignment = Alignment(vertical='center', horizontal='center')
        # Apply bottom border to all cells in merged range
        double_bottom = Border(bottom=Side(style='double'))
        for col in range(1, 6):
            worksheet.cell(row=row, column=col).border = double_bottom
        worksheet.row_dimensions[row].height = 38

        row += 1
        # 日期
        worksheet.merge_cells(f'A{row}:E{row}')
        date_cell = worksheet[f'A{row}']
        date_obj = datetime.fromtimestamp(items[0]['date'])
        date_cell.value = date_obj.strftime('%Y年%m月%d日')
        date_cell.alignment = Alignment(vertical='center', horizontal='center')
        worksheet.row_dimensions[row].height = 30

        row += 1
        # 供应者
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.value = f"供应者：{items[0]['sell_company']}"
        # Apply border to all cells in merged range (A-E)
        for col in range(1, 6):
            set_wrap_border(worksheet.cell(row=row, column=col))
        cell.alignment = Alignment(vertical='center', horizontal='left')
        worksheet.row_dimensions[row].height = 20

        row += 1
        # 表头
        headers = ['材料名称', '规格', '数量', '单位', '备注']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row, column=col)
            cell.value = header
            set_wrap_border(cell)
        worksheet.row_dimensions[row].height = 20

        # 数据行
        for product in items:
            row += 1
            cell_a = worksheet.cell(row=row, column=1)
            cell_a.value = product['product']
            set_wrap_border(cell_a)

            cell_b = worksheet.cell(row=row, column=2)
            cell_b.value = product.get('specification', '')
            set_wrap_border(cell_b)

            cell_c = worksheet.cell(row=row, column=3)
            if product['unit'] in FLOAT_UNITS:
                cell_c.value = round(product['count'], 3)
            else:
                cell_c.value = int(product['count'])
            set_wrap_border(cell_c)

            cell_d = worksheet.cell(row=row, column=4)
            cell_d.value = product['unit']
            set_wrap_border(cell_d)

            cell_e = worksheet.cell(row=row, column=5)
            cell_e.value = ''
            set_wrap_border(cell_e)

            worksheet.row_dimensions[row].height = 20

        # 填充空行
        for _ in range(len(items), 7):
            row += 1
            for col in range(1, 6):
                set_wrap_border(worksheet.cell(row=row, column=col))
            worksheet.row_dimensions[row].height = 20

        # 记账人
        row += 1
        worksheet.merge_cells(f'A{row}:E{row}')
        cell = worksheet[f'A{row}']
        cell.alignment = Alignment(vertical='center', horizontal='right')
        cell.value = f"记账：陈{' ' * 20}"
        worksheet.row_dimensions[row].height = 24

        if index % 2 == 0:
            row += 11
        else:
            row += 3

    # 设置右侧列宽
    worksheet.column_dimensions['H'].width = 20
    worksheet.column_dimensions['I'].width = 18.17
    worksheet.column_dimensions['J'].width = 20
    worksheet.column_dimensions['K'].width = 4.33
    worksheet.column_dimensions['L'].width = 20


def format_data(slim_data, issuing):
    """格式化数据"""
    company_splitted = merge_by_company(slim_data)
    date_splitted = split_by_date(company_splitted)
    count_merged = merge_counts(date_splitted)
    count_splitted = split_by_count(count_merged)

    date_rewritten = rewrite_date(count_splitted, issuing)
    date_sorted = sort_by_date(date_rewritten)

    return date_sorted


def merge_by_company(data):
    """按公司合并"""
    result = {}
    for item in data:
        company = item['sell_company']
        if company in result:
            result[company].append(item)
        else:
            result[company] = [item]
    return list(result.values())


def split_by_date(data):
    """按日期拆分"""
    # 先按日期排序
    for items in data:
        items.sort(key=lambda x: x['date'])

    result = []
    for items in data:
        date_map = {}
        for item in items:
            date_key = item['date']
            if date_key in date_map:
                date_map[date_key].append(item)
            else:
                date_map[date_key] = [item]
        result.extend(list(date_map.values()))

    return result


def sort_by_date(data):
    """按日期排序"""
    # 过滤空列表
    non_empty = [x for x in data if x]
    return sorted(non_empty, key=lambda x: x[0]['date'])


def merge_counts(data):
    """合并数量"""
    result = []
    for items in data:
        # 按产品名称排序（中文）
        try:
            locale.setlocale(locale.LC_COLLATE, 'zh_CN.UTF-8')
            items.sort(key=lambda x: locale.strxfrm(x['product']))
        except:
            items.sort(key=lambda x: x['product'])

        product_map = {}
        for item in items:
            key = f"{item['product']}_{item['unit']}"
            if key in product_map:
                product_map[key]['count'] += item['count']
            else:
                product_map[key] = item.copy()
        result.append(list(product_map.values()))

    return result


def split_by_count(data):
    """按数量拆分（每7条一组）"""
    result = []
    for items in data:
        count = (len(items) + 6) // 7
        for i in range(count):
            result.append(items[i * 7:(i + 1) * 7])
    return result


def rewrite_date(data, issuing):
    """重写日期"""
    # 过滤空列表
    issuing = [x for x in issuing if x]
    if not issuing or not data:
        return data

    first_date = datetime.fromtimestamp(issuing[0][0]['date'])
    # 月初1号
    start = int(first_date.replace(day=1, hour=0, minute=0, second=0).timestamp())
    # first_date 前一天（确保在月初之后）
    last_date = first_date - timedelta(days=1)
    if last_date.month != first_date.month:
        # 如果前一天是上个月，使用当月1号作为结束
        last_date = first_date.replace(day=1)
    last_unix = int(last_date.replace(hour=23, minute=59, second=59).timestamp())

    # 确保 start <= last_unix
    if start > last_unix:
        start, last_unix = last_unix, start
    if start == last_unix:
        last_unix = start + 1

    for items in data:
        for item in items:
            item['date'] = random_range(start, last_unix)

    return data
