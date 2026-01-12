from datetime import datetime
from openpyxl import load_workbook
from .config import INVALID_PRODUCT_TYPES


def safe_float(value):
    """安全地将值转换为浮点数，跳过标签/文本"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        # 跳过包含中文或标签的文本
        if any(char in value for char in [':', '：', '项', '和', '求']):
            return 0
        try:
            return float(value)
        except ValueError:
            return 0
    return 0


def handle_inbound_data(file_path):
    """处理入库发票数据"""
    workbook = load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        data.append(list(row))

    return wash_data(data)


def wash_data(data):
    """清洗数据"""
    slim_data = []

    for item in data[1:]:  # 跳过表头
        if not item or len(item) == 0:
            continue

        # 安全获取列值
        def get_col(idx):
            return item[idx] if len(item) > idx else None

        sell_company = str(get_col(5) or '').strip()

        # 解析产品名称
        product_str = str(get_col(11) or '').strip()
        product_parts = product_str.split('*')
        product_type = product_parts[1] if len(product_parts) > 1 else ''
        product = product_parts[2] if len(product_parts) > 2 else ''

        specification = str(get_col(12) or '').strip()
        unit = str(get_col(13) or '').strip()

        # 解析日期
        date_val = get_col(8)
        if isinstance(date_val, datetime):
            date_unix = int(date_val.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())
        elif date_val:
            try:
                date_obj = datetime.strptime(str(date_val).strip(), '%Y-%m-%d')
                date_unix = int(date_obj.timestamp())
            except:
                date_unix = 0
        else:
            date_unix = 0

        count = safe_float(get_col(14))
        price = safe_float(get_col(16))
        tax = safe_float(get_col(18))

        if product_type not in INVALID_PRODUCT_TYPES and count > 0:
            slim_data.append({
                'sell_company': sell_company,
                'product': product,
                'product_type': product_type,
                'specification': specification,
                'unit': unit,
                'date': date_unix,
                'count': count,
                'price': price,
                'tax': tax
            })

    # 按日期排序
    slim_data.sort(key=lambda x: x['date'])

    return {'valid_data': slim_data}
