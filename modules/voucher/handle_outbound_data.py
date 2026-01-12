import re
from datetime import datetime
from openpyxl import load_workbook
from .config import INVALID_PRODUCT_TYPES


def safe_float(value):
    """安全地将值转换为浮点数，跳过标签/文本"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        # 跳过包含中文或标签的文本
        if any(char in value for char in [':', '：', '项', '和', '求']):
            return 0
        try:
            return float(value)
        except ValueError:
            return 0
    return 0


def handle_outbound_data(file_path):
    """处理出库发票数据"""
    workbook = load_workbook(file_path)
    # 使用第一个工作表（与TypeScript原版一致）
    sheet = workbook[workbook.sheetnames[0]]

    data = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        data.append(list(row))

    return wash_data(data)


def wash_data(data):
    """清洗数据"""
    slim_data = []

    for item in data[1:]:  # 跳过表头
        if not item or len(item) == 0:
            continue

        # 安全获取列值
        def get_col(idx):
            return item[idx] if len(item) > idx else None

        code = str(get_col(3) or '').strip()
        sell_company = str(get_col(5) or '').strip()
        buy_company = str(get_col(7) or '').strip()

        # 解析日期
        date_val = get_col(8)
        if isinstance(date_val, datetime):
            date_unix = int(date_val.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())
        elif date_val:
            date_str = str(date_val).strip()
            date_unix = 0
            # 尝试多种日期格式
            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    date_obj = datetime.strptime(date_str, fmt)
                    date_unix = int(date_obj.replace(hour=0, minute=0, second=0, microsecond=0).timestamp())
                    break
                except ValueError:
                    continue
        else:
            date_unix = 0

        # 解析产品名称
        product_str = str(get_col(11) or '').strip()
        product_parts = product_str.split('*')
        product_type = product_parts[1] if len(product_parts) > 1 else ''
        product_raw = product_parts[2] if len(product_parts) > 2 else ''

        # 提取产品名称
        product_match = re.match(r'([a-zA-Z0-9\-+\u4e00-\u9fa5]+)', product_raw)
        product = product_match.group(1) if product_match else ''

        unit = str(get_col(13) or '').strip()
        count = safe_float(get_col(14))
        notes = str(get_col(26) or '').strip()
        price = safe_float(get_col(16))
        tax = safe_float(get_col(18))

        slim_data.append({
            'code': code,
            'sell_company': sell_company,
            'buy_company': buy_company,
            'date': date_unix,
            'product': product,
            'product_type': product_type,
            'unit': unit,
            'count': count,
            'notes': notes,
            'price': price,
            'tax': tax
        })

    # 找出无效的发票代码
    invalid_codes = []
    for item in slim_data:
        code_match = re.search(r'(\d+)', item['notes'])
        if '被红冲蓝字' in item['notes'] and code_match:
            invalid_codes.append(code_match.group(0))

    valid_data = []
    invalid_data = []

    for item in slim_data:
        is_invalid_type = item['product_type'] in INVALID_PRODUCT_TYPES
        is_reversed = '被红冲蓝字' in item['notes']
        is_invalid_code = item['code'] in invalid_codes

        if is_invalid_type or is_reversed or is_invalid_code:
            if is_reversed:
                code_match = re.search(r'(\d+)', item['notes'])
                matched_code = code_match.group(0) if code_match else ''
                has_matching = any(
                    obj['code'] and obj['code'] == matched_code
                    for obj in slim_data
                )
                if not has_matching:
                    invalid_data.append(item)
        else:
            valid_data.append(item)

    return {'valid_data': valid_data, 'invalid_data': invalid_data}
